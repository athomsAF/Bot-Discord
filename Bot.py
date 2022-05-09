import Audio
import discord
from discord.ext import commands
import youtube_dl
import asyncio
from openpyxl import load_workbook

intents = discord.Intents.default()
intents.members = True
bot = commands.Bot(command_prefix ="$", description = "On découvre python c nice",intents=intents)
ytdl = youtube_dl.YoutubeDL()

dicplay={}
bot.load_extension ("Audio")
################################################################################################################
#############################################Démarrage##########################################################
################################################################################################################
@bot.event
async def on_ready():
    print("--- Ready ---")
    await bot.change_presence(activity=discord.Activity(type=discord.ActivityType.watching, name="Netflix"))
    channel = bot.get_channel(963007190308892702)
    await channel.send("< ON >",delete_after=10)
    wbform= load_workbook('formation.xlsx')
    wsform=wbform.active
    (wbplay,wsplay)=loadPlaylistExcel()

    



################################################################################################################
######################################Ensemble des non-commandes################################################
################################################################################################################

def loadPlaylistExcel():
    wbplay= load_workbook('playlist.xlsx')
    wsplay=wbplay["NomId"]
    wbplay.save('playlist.xlsx')
    return(wbplay,wsplay)

def searchName(wsplay):
    nbLigne=wsplay.max_row
    Names=[]
    for i in range(2,nbLigne+1):
        if wsplay[f"A{i}"].value!=None:
            Names+=[wsplay[f"A{i}"].value]
    if Names==[]:
        Names=0
    print(Names)
    return(Names)

def searchPlaylist(name,wbplay):
    wsplayA=wbplay[f"{name}playlist"]
    nbLigne=wsplayA.max_row
    PlaylistNames=[]
    for i in range(1,nbLigne+1):
        if wsplayA[f"A{i}"].value !=None:
            PlaylistNames+=[wsplayA[f"A{i}"].value]
    return(PlaylistNames)

def searchPlaylistList(name,plname,wbplay):
    wsplayA=wbplay[f"{name}playlist"]
    position=searchPlaylist(name,wbplay).index(plname)
    i=66
    listePlaylist=[]
    while i>0:
        if  wsplayA[f"{chr(i)}{position+1}"].value!= None:
            listePlaylist+=[wsplayA[f"{chr(i)}{position+1}"].value]
            i+=1
        else:
            i=0
    return(listePlaylist)

def searchVideoName (link):
    if link[0:4]=="http":
        video = ytdl.extract_info(link, download=False)
    else:
        video = ytdl.extract_info("ytsearch:%s" % link, download=False)['entries'][0]
    return(video['title'])





################################################################################################################
######################################ENSEMBLE DE COMMANDES#####################################################
################################################################################################################
#add music playlistc
#@bot.command()
#async def refresh(self,ctx,wbplay,wsplay):
#    members=await ctx.guild.fetch_members(limit=150).flatten()
#    j=0
#    for i in range (len(members)):
#        if members[j].name=="Klyde" or members[j].name=="FabLaBot" :
#            del members[j]
#            j-=1
#        j+=1
#
#    for i in range(len(members)):  
#            wsplay['A1']="Name"
#            wsplay['B1']="Id"
#            wsplay[f"A{i+2}"]=members[i].name
#            wsplay[f"B{i+2}"]=members[i].id
#            wsplay[f"C{i+2}"]=i+1
#            if wbplay.sheetnames.count(f"{members[i].name}playlist")==0:
#                wbplay.create_sheet(f"{members[i].name}playlist")
#            elif wbplay.sheetnames.count(f"{members[i].name}playlist")>1:
#                for j in [i for i, e in enumerate(wbplay.sheetnames) if e == f"{members[i].name}playlist"]:
#                    wbplay.remove_sheet(wbplay[wbplay.sheetnames[j]])
#                wbplay.create_sheet(f"{members[i].name}playlist")
#        elif wbplay.sheetnames.count(f"{members[i].name}playlist")>0:
#            for j in range (wbplay.sheetnames.count(f"{members[i].name}playlist")):
#                wbplay.remove(wbplay[f"{members[i].name}playlist"])
#    print(wbplay.sheetnames)
#    wbplay.save('playlist.xlsx')

#####################################################COGS#########################################################################

@bot.command()
async def reload(ctx, name=None):
    if name:
        try:
            bot.reload_extension(name)
        except:
            bot.load_extension(name)

###############################################################################################################################



@bot.command()
async def add(ctx,*reason):
    await ctx.message.delete()
    link=str(" ".join(reason))
    if link[0:4]=="http":
        link = searchVideoName(link)
    (wbplay,wsplay)=loadPlaylistExcel()
    wsp = wbplay[f"{ctx.message.author.name}playlist"]
    if wsp['A1'].value!=None:
        liste=[]
        i=0
        while i!=-1:
            i+=1
            if (wsp[f"A{i}"].value!=None):
                liste+=[wsp[f"A{i}"].value]
            else:
                i=-1
        var=str("\n-".join(liste))
        embed= discord.Embed(
            title='Quel playlist voulez-vous choisir (tapez le numéro)',
            description=f"Choisissez : \n-{var} \n-nouvelle playlist (taper le nom) \n\n|| Cette requête s'arrêtera dans 30 secondes ||"
        )
        sent= await ctx.send (embed=embed)
        
        try:
            msg= await bot.wait_for("message",timeout=30,check=lambda message: message.author == ctx.author and message.channel==ctx.channel)   
            if msg:
                await sent.delete()
                await msg.delete()

        except asyncio.TimeoutError: 
            await sent.delete()
            await ctx.send("Annulation du au TimeOut", delete_after=10)

        if liste.count(str(msg.content))==0:
            wsp[f"A{len(liste)+1}"].value=msg.content
            wsp[f"B{len(liste)+1}"].value=link
        elif  liste.count(str(msg.content))>0:
            j=65
            while j!=0:
                j+=1
                if j>91:
                    if (wsp[f"{chr((j-64)//26+64)}{chr((j-64)%26+64)}{liste.count(str(msg.content))+1}"].value==None):
                        wsp[f"{chr((j-64)//26+64)}{chr((j-64)%26+64)}{liste.count(str(msg.content))+1}"].value=link
                        j=0
                else:
                    if(wsp[f"{chr(j)}{liste.index(str(msg.content))+1}"].value==None):
                        wsp[f"{chr(j)}{liste.index(str(msg.content))+1}"].value=link
                        j=0
        liste=searchPlaylistList(ctx.message.author.name,msg.content,wbplay)
        var="-"+str("\n-".join(liste))
    else:
        embed= discord.Embed(
            title='Choisissez un titre de playlist',
            description="||Cette requête s'arrêtera dans 30 secondes||"
        )
        sent= await ctx.send (embed=embed)

        try:
            msg= await bot.wait_for("message",timeout=30,check=lambda message: message.author == ctx.author and message.channel==ctx.channel)   
            if msg:
                await sent.delete()
                await msg.delete()
                wsp['A1'].value=msg.content
                wsp['B1'].value=link
                var="-" +link
        except asyncio.TimeoutError: 
            await sent.delete()
            await ctx.send("Annulation du au TimeOut", delete_after=10)

    ##########Affichage final###############
    
    embed= discord.Embed(
        title='Voici donc votre playlist',
        description=f"Votre playlist: **{msg.content}** est composé de:\n*{var}*\n\n|| Cette requête s'arrêtera dans 10 secondes ||"
    )
    sent=await ctx.send (embed=embed,delete_after=10)
    wbplay.save('playlist.xlsx')


@bot.command()
async def playlist(ctx):
    (wbplay,wsplay)=loadPlaylistExcel()
    await ctx.send("- "+str("\n-".join(searchPlaylist(ctx.message.author.name,wbplay))),delete_after=20)

#Commande de CLEAR
@bot.command(aliases= ['clear','Cl','CL','cl']) #clear command
@commands.has_permissions(manage_messages=True)
async def Clear(ctx,amount: int = None):
    if amount == None:
        await ctx.channel.purge(limit=1000000)
    else:
       await ctx.channel.purge(limit=amount+1)

#Commande infoserveur
@bot.command(aliases= ['Infoserveur','Info','info'])
async def InfoServeur(ctx):
    server = ctx.guild
    NombreChanText = len(server.text_channels)
    NombreChanVoc = len(server.voice_channels)
    NombrePersonnes = server.member_count
    NomServeur = server.name
    message = f"Le serveur **{NomServeur}** contient {NombrePersonnes} personnes. \nLe serveur possède {NombreChanText} salons textuels et {NombreChanVoc} salons vocaux."
    await ctx.send(message)

#Commande de Kick
@bot.command()
async def kick(ctx, user : discord.User, *reason):
    reason = " ".join(reason)
    await ctx.guild.kick(user , reason = reason)
    await ctx.send(f"{user} a été kick.")


    @bot.command(aliases= ['hi'])
    async def Hi(ctx):
        await ctx.send(f"Yoosh {ctx.author.mention} !", )

#@bot.command()
#async def Poll(ctx, question, options, *choice):
# 
#    message = f"{question} \n"
#    for i in range(int(options)):
#        message += choice,"\n"
#    
#    await ctx.send(message)





#   /usr/bin/python3 "/Volumes/Macintosh HD - Données/FICHIERS/ESILV/VsCode/Bot-Discord/Bot.py"
#


bot.run("OTYzMzg1MTY3MjgyNTg5NzA3.YlVUWg.DAEkTP7NxDCOdSbbd6n6Qpg9o-U")